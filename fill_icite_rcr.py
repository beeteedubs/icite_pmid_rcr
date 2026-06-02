#!/usr/bin/env python3
"""Fill Mean RCR (col M) and Weighted RCR (col N) from iCite API using PMIDs in col E."""

import argparse
import re
import shutil
import time
from pathlib import Path

import openpyxl
import requests

ICITE_API = "https://icite.od.nih.gov/api/pubs"
# iCite allows up to 1000 PMIDs per call; response size limits ~350 PMIDs with fl filter.
BATCH_SIZE = 300
REQUEST_DELAY = 0.2

DEFAULT_INPUT = Path(__file__).parent / "ABA Anesthesiologit Data Extracted.xlsx"
DEFAULT_SHEET = "Pain Clean"
SHEET_SUFFIX = " Clean"

COL_PMIDS = 5  # E
COL_MEAN_RCR = 13  # M
COL_WEIGHTED_RCR = 14  # N


def parse_pmids(cell_value) -> list[str]:
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    return [p for p in re.split(r"\s+", text) if p.isdigit()]


def specialty_sheets(sheetnames: list[str]) -> list[str]:
    return [name for name in sheetnames if name.endswith(SHEET_SUFFIX)]


def fetch_rcr_cache(pmids: set[str], session: requests.Session) -> dict[str, float]:
    cache: dict[str, float] = {}
    sorted_pmids = sorted(pmids, key=int)
    total_batches = (len(sorted_pmids) + BATCH_SIZE - 1) // BATCH_SIZE

    for i in range(0, len(sorted_pmids), BATCH_SIZE):
        batch = sorted_pmids[i : i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  API batch {batch_num}/{total_batches} ({len(batch)} PMIDs)...")

        resp = session.get(
            ICITE_API,
            params={
                "pmids": ",".join(batch),
                "fl": ["pmid", "relative_citation_ratio"],
            },
            timeout=120,
        )
        resp.raise_for_status()
        for row in resp.json().get("data", []):
            cache[str(row["pmid"])] = float(row.get("relative_citation_ratio") or 0.0)

        if i + BATCH_SIZE < len(sorted_pmids):
            time.sleep(REQUEST_DELAY)

    return cache


def compute_metrics(pmids: list[str], cache: dict[str, float]) -> tuple[float, float]:
    rcrs = [cache.get(p, 0.0) for p in pmids]
    weighted = round(sum(rcrs), 2)
    mean = round(weighted / len(pmids), 2)
    return mean, weighted


def process_sheet(ws, cache: dict[str, float]) -> int:
    rows_with_pmids: list[tuple[int, list[str]]] = []

    for row_idx in range(2, ws.max_row + 1):
        pmids = parse_pmids(ws.cell(row=row_idx, column=COL_PMIDS).value)
        if pmids:
            rows_with_pmids.append((row_idx, pmids))

    print(f"  {ws.title}: {len(rows_with_pmids)} rows with PMIDs")

    for row_idx, pmids in rows_with_pmids:
        mean, weighted = compute_metrics(pmids, cache)
        ws.cell(row=row_idx, column=COL_MEAN_RCR, value=mean)
        ws.cell(row=row_idx, column=COL_WEIGHTED_RCR, value=weighted)

    return len(rows_with_pmids)


def collect_pmids(ws) -> set[str]:
    pmids: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        pmids.update(parse_pmids(ws.cell(row=row_idx, column=COL_PMIDS).value))
    return pmids


def main():
    parser = argparse.ArgumentParser(description="Fill iCite RCR metrics in Excel workbook")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help=f"Single sheet to process (default: {DEFAULT_SHEET})")
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help=f"Process all sheets ending with '{SHEET_SUFFIX.strip()}'",
    )
    parser.add_argument("--dry-run", action="store_true", help="Fetch and compute but do not save")
    parser.add_argument("--no-backup", action="store_true", help="Skip creating .bak before save")
    args = parser.parse_args()

    input_path = args.input.resolve()
    if not input_path.exists():
        raise SystemExit(f"File not found: {input_path}")

    wb = openpyxl.load_workbook(input_path, read_only=False)

    if args.all_sheets:
        sheets_to_run = specialty_sheets(wb.sheetnames)
        if not sheets_to_run:
            wb.close()
            raise SystemExit(f"No sheets ending with '{SHEET_SUFFIX.strip()}' found.")
    else:
        sheet_name = args.sheet or DEFAULT_SHEET
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise SystemExit(f"Sheet not found: {sheet_name!r}. Available: {wb.sheetnames}")
        sheets_to_run = [sheet_name]

    print(f"Loading {input_path}")
    print(f"Sheets: {', '.join(sheets_to_run)}")

    all_pmids: set[str] = set()
    for sheet_name in sheets_to_run:
        all_pmids.update(collect_pmids(wb[sheet_name]))

    print(f"Total unique PMIDs across sheet(s): {len(all_pmids)}")
    print("Fetching RCR values from iCite API (needs internet)...")
    session = requests.Session()
    cache = fetch_rcr_cache(all_pmids, session)
    found = sum(1 for p in all_pmids if p in cache)
    print(f"  Retrieved RCR for {found}/{len(all_pmids)} unique PMIDs")

    total_rows = 0
    for sheet_name in sheets_to_run:
        total_rows += process_sheet(wb[sheet_name], cache)

    print(f"Updated {total_rows} rows total (columns M=Mean RCR, N=Weighted RCR)")

    if args.dry_run:
        print("Dry run — not saving.")
        wb.close()
        return

    if not args.no_backup:
        backup_path = input_path.with_suffix(input_path.suffix + ".bak")
        if not backup_path.exists():
            print(f"Creating backup: {backup_path}")
            shutil.copy2(input_path, backup_path)

    print(f"Saving {input_path}...")
    wb.save(input_path)
    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
